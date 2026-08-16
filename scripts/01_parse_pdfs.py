import csv
import datetime as dt
import math
import os
import pathlib
import re
import argparse

import pdfplumber


ROOT = pathlib.Path(__file__).resolve().parent.parent
DEFAULT_INPUT_DIR = ROOT / "data" / "input"
DEFAULT_OCR_FILE = ROOT / "outputs" / "ocr.txt"
DEFAULT_OUTPUT_DIR = ROOT / "outputs"


def clean_text(value):
    if value is None:
        return ""
    value = str(value).replace("\n", "")
    return re.sub(r"\s+", " ", value).strip()


def parse_pdf_header(text):
    period = re.search(r"行程起止日期[：:]\s*(\d{4}-\d{2}-\d{2})\s*至\s*(\d{4}-\d{2}-\d{2})", text)
    count = re.search(r"共\s*(\d+)\s*笔行程", text)
    total = re.search(r"合计\s*([\d.]+)\s*元", text)
    return {
        "start": dt.date.fromisoformat(period.group(1)) if period else None,
        "end": dt.date.fromisoformat(period.group(2)) if period else None,
        "count": int(count.group(1)) if count else None,
        "total": float(total.group(1)) if total else None,
    }


def normalize_header(cell):
    return clean_text(cell)


def map_columns(header):
    normalized = [normalize_header(cell) for cell in header]
    mapping = {}
    for i, name in enumerate(normalized):
        if name == "序号":
            mapping["seq"] = i
        elif name == "车型类别":
            mapping["category"] = i
        elif name == "车型" and "category" not in mapping or name == "车型" and mapping.get("车型") is None:
            mapping.setdefault("vehicle", i)
        elif "上车时间" in name:
            mapping["depart"] = i
        elif "出发时间" in name:
            mapping["depart"] = i
        elif "到达时间" in name:
            mapping["arrive"] = i
        elif name == "城市":
            mapping["city"] = i
        elif name == "起点":
            mapping["origin"] = i
        elif name == "终点":
            mapping["dest"] = i
        elif "里程" in name:
            mapping["mileage"] = i
        elif "金额" in name:
            mapping["amount"] = i
        elif "备注" in name:
            mapping["remarks"] = i
    # Fix any ambiguous 车型 assignment.
    for i, name in enumerate(normalized):
        if name == "车型" and "vehicle" not in mapping:
            mapping["vehicle"] = i
    return mapping


def parse_mm_dd_time(value):
    m = re.search(r"(\d{2})-(\d{2})\s+(\d{2}):(\d{2})", value or "")
    if not m:
        return None
    return int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))


def parse_date_token(value):
    m = re.search(r"(\d{2})-(\d{2})\s+(\d{2}):(\d{2})", value or "")
    if not m:
        return None
    return int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))


def infer_years(rows, start, end):
    previous = None
    for row in rows:
        parts = parse_mm_dd_time(row.get("depart") or row.get("depart_time") or "")
        if not parts:
            row["start_time"] = ""
            row["end_time"] = ""
            continue
        month, day, hour, minute = parts
        chosen = None
        for year in range(start.year, end.year + 1):
            candidate = dt.datetime(year, month, day, hour, minute)
            candidate_date = candidate.date()
            if candidate_date < start or candidate_date > end:
                continue
            if previous is not None and candidate < previous:
                continue
            chosen = candidate
            break
        if chosen is None:
            chosen = dt.datetime(start.year, month, day, hour, minute)
        row["start_time"] = chosen.strftime("%Y-%m-%d %H:%M:%S")
        previous = chosen

        if "arrive" in row and row["arrive"]:
            arrive_parts = parse_date_token(row["arrive"])
            if arrive_parts:
                amonth, aday, ahour, aminute = arrive_parts
                arrive_dt = dt.datetime(
                    chosen.year, amonth, aday, ahour, aminute
                )
                # A cross-midnight hitch should stay on the same day or the next day.
                if arrive_dt < chosen:
                    arrive_dt = dt.datetime(chosen.year + 1, amonth, aday, ahour, aminute)
                row["end_time"] = arrive_dt.strftime("%Y-%m-%d %H:%M:%S")


def extract_pdf_rows(path):
    with pdfplumber.open(path) as pdf:
        text = pdf.pages[0].extract_text() or ""
        meta = parse_pdf_header(text)
        rows = []
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                if len(table) < 2:
                    continue
                header = table[0]
                mapping = map_columns(header)
                for raw in table[1:]:
                    if not raw or not raw[0] or not clean_text(raw[0]).isdigit():
                        continue
                    row = {}
                    for field, idx in mapping.items():
                        row[field] = clean_text(raw[idx]) if idx < len(raw) else ""
                    rows.append(row)

    infer_years(rows, meta["start"], meta["end"])
    return meta, rows


def numeric(value):
    if value in (None, ""):
        return ""
    value = str(value).replace(",", "").replace("¥", "").strip()
    try:
        return float(value)
    except ValueError:
        return ""


def standardize_pdf_rows(meta, rows, filename):
    out = []
    for row in rows:
        out.append({
            "来源": "行程单PDF",
            "来源文件": filename,
            "来源序号": row.get("seq", ""),
            "车型": row.get("vehicle", ""),
            "车型类别": row.get("category", ""),
            "上车时间": row.get("start_time", ""),
            "到达时间": row.get("end_time", ""),
            "城市": row.get("city", ""),
            "起点": row.get("origin", ""),
            "终点": row.get("dest", ""),
            "里程_公里": numeric(row.get("mileage", "")),
            "金额_元": numeric(row.get("amount", "")),
            "备注": row.get("remarks", ""),
            "企业支付": "",
            "个人实付_元": "",
            "订单金额_元": "",
        })
    return out


def parse_ocr_records(ocr_file):
    records = []
    current_file = ""
    current = None
    bullet_index = 0

    for line in ocr_file.read_text(encoding="utf-8").splitlines():
        if line.startswith("===== "):
            if current and current.get("file"):
                records.append(current)
            current_file = line.strip("= ").strip()
            current = None
            bullet_index = 0
            continue
        parts = line.split("\t", 4)
        if len(parts) != 5:
            continue
        text = parts[4].strip()

        date_match = re.search(
            r"专车\s*(\d{4})\.(\d{2})\.(\d{2})\s*(?:\S+\s*)?(\d{2}):(\d{2})",
            text,
        )
        if date_match:
            if current and current.get("file"):
                records.append(current)
            y, m, d, hh, mm = date_match.groups()
            current = {
                "file": current_file,
                "date": f"{y}-{m}-{d}",
                "time": f"{hh}:{mm}:00",
                "origin": "",
                "dest": "",
                "amount": "",
                "enterprise": "是",
                "personal": "0",
            }
            bullet_index = 0
            continue

        if current is None:
            continue

        if text.startswith("•") or text.startswith("·") or text.startswith("."):
            place = re.sub(r"^[•·.\s]+", "", text).strip()
            if not place:
                continue
            if bullet_index == 0:
                current["origin"] = place
            elif bullet_index == 1:
                current["dest"] = place
            bullet_index += 1
            continue

        amount_match = re.search(
            r"订[^\d]{0,4}额[^\d]{0,4}¥?\s*(\d+(?:\.\d+)?)", text
        )
        if amount_match:
            if not current.get("amount"):
                current["amount"] = amount_match.group(1)
            continue

        if "企业支付" in text:
            current["enterprise"] = "是"
            continue

        if re.search(r"¥\s*0\b", text):
            current["personal"] = "0"
            continue

    if current and current.get("file"):
        records.append(current)

    # Merge the same trip shown at the bottom of one screenshot and the top of the next.
    by_key = {}
    for rec in records:
        key = (rec["date"], rec["time"])
        if key not in by_key:
            by_key[key] = rec
            continue
        old = by_key[key]
        other = rec
        score = lambda r: sum(bool(r.get(f)) for f in ("origin", "dest", "amount"))
        if score(other) > score(old):
            old, other = other, old
        for field in ("origin", "dest", "amount"):
            if not old.get(field) and other.get(field):
                old[field] = other[field]
        by_key[key] = old

    result = sorted(by_key.values(), key=lambda r: (r["date"], r["time"]))
    return result


def standardize_screenshot_rows(records):
    out = []
    for i, rec in enumerate(records, 1):
        full_time = f"{rec['date']} {rec['time']}"
        incomplete = []
        origin = re.sub(r"(?:…|\.+|•|·|．)+$", "", rec.get("origin", "")).strip()
        dest = re.sub(r"(?:…|\.+|•|·|．)+$", "", rec.get("dest", "")).strip()
        if not origin:
            incomplete.append("起点未识别")
        if not dest:
            incomplete.append("终点未识别")
        if not rec.get("amount"):
            incomplete.append("金额未识别")
        out.append({
            "来源": "行程截图",
            "来源文件": os.path.basename(rec["file"]),
            "来源序号": i,
            "车型": "专车",
            "车型类别": "",
            "上车时间": full_time,
            "到达时间": "",
            "城市": "北京",
            "起点": origin,
            "终点": dest,
            "里程_公里": "",
            "金额_元": numeric(rec.get("amount", "")),
            "备注": "；".join(incomplete),
            "企业支付": rec.get("enterprise", "是"),
            "个人实付_元": numeric(rec.get("personal", "")),
            "订单金额_元": numeric(rec.get("amount", "")),
        })
    return out


def main():
    parser = argparse.ArgumentParser(
        description="Parse Didi trip receipt PDFs and optional screenshot OCR text."
    )
    parser.add_argument("--input-dir", type=pathlib.Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--ocr-file", type=pathlib.Path, default=DEFAULT_OCR_FILE)
    parser.add_argument("--out", type=pathlib.Path, default=DEFAULT_OUTPUT_DIR / "trips.csv")
    parser.add_argument("--xlsx", type=pathlib.Path, default=DEFAULT_OUTPUT_DIR / "trips.xlsx")
    args = parser.parse_args()

    args.out.parent.mkdir(parents=True, exist_ok=True)
    pdf_files = sorted(args.input_dir.glob("滴滴出行行程报销单*.pdf"))
    extra = args.input_dir / "滴滴顺风车行程单.pdf"
    if extra.exists():
        pdf_files.append(extra)

    all_rows = []
    source_summary = []
    for path in pdf_files:
        meta, rows = extract_pdf_rows(path)
        std = standardize_pdf_rows(meta, rows, path.name)
        all_rows.extend(std)
        source_summary.append({
            "来源文件": path.name,
            "记录数": len(std),
            "合计金额": meta["total"],
            "起止日期": f"{meta['start']} 至 {meta['end']}",
        })

    screenshot_records = []
    screenshot_rows = []
    if args.ocr_file.exists():
        screenshot_records = parse_ocr_records(args.ocr_file)
        screenshot_rows = standardize_screenshot_rows(screenshot_records)
    else:
        print(f"Skipping screenshots: {args.ocr_file} not found")
    all_rows.extend(screenshot_rows)

    fields = [
        "来源", "来源文件", "来源序号", "车型", "车型类别", "上车时间", "到达时间",
        "城市", "起点", "终点", "里程_公里", "金额_元", "备注", "企业支付",
        "个人实付_元", "订单金额_元",
    ]

    csv_path = args.out
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_rows)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "行程明细"
    ws.append(fields)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    for row in all_rows:
        ws.append([row.get(field, "") for field in fields])
    for col_idx in range(1, len(fields) + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, min(60, len(str(cell.value))))
        ws.column_dimensions[letter].width = max_len + 2
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{ws.max_row}"

    ws2 = wb.create_sheet("来源概览")
    ws2.append(["来源文件", "记录数", "合计金额", "起止日期"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    for item in source_summary:
        ws2.append([item["来源文件"], item["记录数"], item["合计金额"], item["起止日期"]])
    if screenshot_records:
        dates = sorted(rec["date"] for rec in screenshot_records)
        screenshot_label = f"行程截图（{len(screenshot_rows)}张）"
        screenshot_range = f"{dates[0]} 至 {dates[-1]}" if dates else ""
    else:
        screenshot_label = "行程截图（无）"
        screenshot_range = ""
    ws2.append([screenshot_label, len(screenshot_rows), "", screenshot_range])
    for col_idx in range(1, 5):
        letter = get_column_letter(col_idx)
        ws2.column_dimensions[letter].width = 42
    ws2.freeze_panes = "A2"

    xlsx_path = args.xlsx
    wb.save(xlsx_path)

    print(f"PDF records: {sum(len(_) for _ in [source_summary]) and len(all_rows) - len(screenshot_rows)}")
    print(f"Screenshot records: {len(screenshot_rows)}")
    print(f"Total records: {len(all_rows)}")
    print(f"CSV: {csv_path}")
    print(f"XLSX: {xlsx_path}")
    print("Screenshot sample:")
    for rec in screenshot_records[:5] + screenshot_records[-5:]:
        print(rec)


if __name__ == "__main__":
    main()
